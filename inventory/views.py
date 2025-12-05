import io
from decimal import Decimal, InvalidOperation

from django.http import HttpResponse
from django.db import transaction
from rest_framework import filters, viewsets, status
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.response import Response

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .models import CatalogItem, ItemCategory
from .serializers import CatalogItemSerializer


# -----------------------------------
# :: CatalogItem ViewSet
# -----------------------------------

class CatalogItemViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing CatalogItems with:
    - Filtering, searching, ordering
    - Excel import/export
    - Template generation for imports
    """

    serializer_class = CatalogItemSerializer
    permission_classes = (IsAuthenticated,)
    filter_backends = (filters.SearchFilter, filters.OrderingFilter)
    search_fields = ("name", "category", "notes")
    ordering_fields = ("name", "category", "updated_at")
    ordering = ("name",)
    parser_classes = (FormParser, MultiPartParser, JSONParser)

    # -----------------------------------
    # :: Queryset with smart filters
    # -----------------------------------

    def get_queryset(self):
        queryset = CatalogItem.objects.all()

        # Filter by active status
        active_param = self.request.query_params.get("active")
        if active_param:
            active_param = active_param.lower()
            if active_param in {"1", "true", "yes", "y"}:
                queryset = queryset.filter(is_active=True)
            elif active_param in {"0", "false", "no", "n"}:
                queryset = queryset.filter(is_active=False)

        # Filter by category
        category = self.request.query_params.get("category")
        if category:
            queryset = queryset.filter(category=category)

        # Optimize related queries if needed (e.g., foreign keys)
        return queryset.select_related()  # Adjust if relations exist

    # -----------------------------------
    # :: Excel Export
    # -----------------------------------

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Export catalog items to Excel with styled headers."""
        wb, ws = self._create_workbook("Catalog Items")
        headers = ['ID', 'Name', 'Category', 'Count Unit', 'Order Unit', 'Pack Size',
                   'Is Active', 'Helper Text', 'Notes']
        self._write_headers(ws, headers)

        # Smart query: only fetch needed fields and order by name
        items = CatalogItem.objects.values(
            'id', 'name', 'category', 'count_unit', 'order_unit',
            'pack_size', 'is_active', 'helper_text', 'notes'
        ).order_by('name')

        self._write_rows(ws, items, headers)
        return self._return_excel_response(wb, "catalog_items.xlsx")

    # -----------------------------------
    # :: Excel Import
    # -----------------------------------

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def import_excel(self, request):
        """Import catalog items from an Excel file."""
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response({'error': 'File must be Excel (.xlsx or .xls)'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
        except Exception as e:
            return Response({'error': f'Cannot read Excel file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        headers = [cell.value for cell in ws[1]]
        required_headers = ['Name', 'Category',
                            'Count Unit', 'Order Unit', 'Pack Size']
        missing_headers = [h for h in required_headers if h not in headers]
        if missing_headers:
            return Response({'error': f'Missing required columns: {", ".join(missing_headers)}'}, status=status.HTTP_400_BAD_REQUEST)

        valid_categories = set(choice[0] for choice in ItemCategory.choices)
        created_count, updated_count, errors = 0, 0, []

        # Smart transaction to reduce DB writes
        with transaction.atomic():
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue  # Skip empty rows
                row_data = dict(zip(headers, row))
                try:
                    name = (row_data.get('Name') or '').strip()
                    if not name:
                        errors.append(f'Row {row_num}: Name is required')
                        continue

                    category = (row_data.get('Category') or 'other').lower()
                    if category not in valid_categories:
                        category = 'other'

                    try:
                        pack_size = Decimal(
                            str(row_data.get('Pack Size') or 1))
                        if pack_size <= 0:
                            pack_size = Decimal('1')
                    except (InvalidOperation, TypeError):
                        pack_size = Decimal('1')

                    is_active_value = row_data.get('Is Active', 'Yes')
                    is_active = str(is_active_value).strip().lower() in [
                        'yes', 'true', '1', 'active']

                    defaults = {
                        'category': category,
                        'count_unit': row_data.get('Count Unit', 'each') or 'each',
                        'order_unit': row_data.get('Order Unit', 'case') or 'case',
                        'pack_size': pack_size,
                        'is_active': is_active,
                        'helper_text': row_data.get('Helper Text', '') or '',
                        'notes': row_data.get('Notes', '') or '',
                    }

                    item, created = CatalogItem.objects.update_or_create(
                        name=name, defaults=defaults)
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1

                except Exception as e:
                    errors.append(f'Row {row_num}: {str(e)}')

        return Response({
            'success': True,
            'created': created_count,
            'updated': updated_count,
            'errors': errors or None
        })

    # -----------------------------------
    # :: Excel Template
    # -----------------------------------

    @action(detail=False, methods=['get'])
    def template(self, request):
        """Generate Excel template for importing catalog items."""
        wb, ws = self._create_workbook("Template")
        headers = ['Name', 'Category', 'Count Unit', 'Order Unit',
                   'Pack Size', 'Is Active', 'Helper Text', 'Notes']
        self._write_headers(ws, headers)

        # Sample row
        sample_row = ['Sample Item', 'fruit', 'bags', 'cases',
                      6, 'Yes', '1 case = 6 bags', 'Sample notes']
        self._write_rows(ws, [sample_row], headers, is_dict=False)

        # Instructions sheet
        instructions = wb.create_sheet("Instructions")
        instructions['A1'] = "Import Instructions"
        instructions['A1'].font = Font(bold=True, size=14)
        instruction_text = [
            "",
            "1. Fill in the Template sheet with your catalog items",
            "2. Required fields: Name, Category, Count Unit, Order Unit, Pack Size",
            "3. Category options: fruit, dairy, dry, packaging, other",
            "4. Is Active: Yes/No or True/False",
            "5. Pack Size must be a positive number",
            "6. Save and upload the file",
            "",
            "Note: Items with the same name will be updated; new names will create new items"
        ]
        for row, text in enumerate(instruction_text, start=2):
            instructions[f'A{row}'] = text
        instructions.column_dimensions['A'].width = 80

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        return self._return_excel_response(wb, "catalog_import_template.xlsx")

    # -----------------------------------
    # :: Helper Methods
    # -----------------------------------

    def _create_workbook(self, title):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title
        return wb, ws

    def _write_headers(self, ws, headers):
        font = Font(bold=True, color="FFFFFF")
        fill = PatternFill(start_color="6366F1",
                           end_color="6366F1", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = font
            cell.fill = fill
            cell.alignment = alignment
            cell.border = border

    def _write_rows(self, ws, data, headers, is_dict=True):
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        for row_num, row in enumerate(data, start=2):
            for col_num, header in enumerate(headers, start=1):
                value = row[header] if is_dict else row[col_num-1]
                ws.cell(row=row_num, column=col_num,
                        value=value).border = border

    def _return_excel_response(self, wb, filename):
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
